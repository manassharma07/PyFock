from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.patches import Patch
from openpyxl import load_workbook


# ==== CONFIG FLAGS ====
x_axis_choice = "water"  # "water" or "basis"
log_scale = False        # True for log scale, False for linear

# Toggle bars here. For example, set PySCF bars to False for a PyFock-only plot,
# or set the 4-core bars to False for a 32-core-only plot.
bar_enabled = {
    "PySCF 4 core": True,
    "PySCF 32 core": True,
    "PyFock 4 core CPU": True,
    "PyFock 32 core CPU": True,
    "PyFock GPU": True,
}

# ==== DATA SOURCE ====
workbook_path = Path(__file__).with_suffix(".xlsx")
water_molecules = np.array([47, 76, 100, 139])

series_specs = [
    {
        "key": "PySCF 4 core",
        "section": ("PySCF", "CPU 4 cores"),
        "label": "PySCF CPU (4 core)",
        "short_label": "PySCF 4c",
        "kind": "pyscf",
        "color": "#9BB7D4",
        "edgecolor": "#1F4E79",
    },
    {
        "key": "PyFock 4 core CPU",
        "section": ("PyFock", "CPU 4 cores"),
        "label": "PyFock CPU (4 core)",
        "short_label": "PyFock CPU 4c",
        "kind": "pyfock",
        "edgecolor": "#6B7280",
        "alpha": 1.0,
    },
    {
        "key": "PySCF 32 core",
        "section": ("PySCF", "CPU 32 cores"),
        "label": "PySCF CPU (32 core)",
        "short_label": "PySCF 32c",
        "kind": "pyscf",
        "color": "#3E6FA3",
        "edgecolor": "#17324D",
    },
    {
        "key": "PyFock 32 core CPU",
        "section": ("PyFock", "CPU 32 cores"),
        "label": "PyFock CPU (32 core)",
        "short_label": "PyFock CPU 32c",
        "kind": "pyfock",
        "edgecolor": "#374151",
        "alpha": 1.0,
    },
    {
        "key": "PyFock GPU",
        "section": ("GPU",),
        "label": "PyFock GPU",
        "short_label": "PyFock GPU",
        "kind": "pyfock",
        "edgecolor": "#111827",
        "alpha": 1.0,
    },
]

pyfock_components = [
    ("J time (s) / iter", "ERI", "#D9902F"),
    ("XC (s) / iter", "XC", "#4B9B82"),
    ("Other", "Other", "#7E8794"),
]


def clean_cell(value):
    return str(value).strip() if value is not None else ""


def load_timing_sections(path):
    ws = load_workbook(path, data_only=True).active
    sections = {}
    current_name = None
    current_machine = None
    headers = None

    for row in ws.iter_rows(values_only=True):
        first = clean_cell(row[0])

        if first in {"GPU", "PyFock", "PySCF"}:
            current_name = first
            current_machine = None
            headers = None
            continue

        if first.startswith("CPU") and current_name in {"PyFock", "PySCF"}:
            current_machine = first
            headers = None
            continue

        if first == "No. of water molecules":
            headers = [clean_cell(value) for value in row]
            continue

        if not isinstance(row[0], (int, float)) or headers is None:
            continue

        section_key = (current_name,) if current_name == "GPU" else (current_name, current_machine)
        values = dict(zip(headers, row))
        water_count = int(row[0])
        sections.setdefault(section_key, {})[water_count] = values

    return sections


def values_for_section(sections, section_key, column_name):
    missing = [water for water in water_molecules if water not in sections.get(section_key, {})]
    if missing:
        raise ValueError(f"Missing water clusters {missing} for section {section_key}")

    values = [
        sections[section_key][int(water)][column_name]
        for water in water_molecules
    ]
    return np.array(values, dtype=float)


sections = load_timing_sections(workbook_path)

series = []
for spec in series_specs:
    if not bar_enabled.get(spec["key"], False):
        continue

    total_times = values_for_section(
        sections,
        spec["section"],
        "Total Time Taken (s) / iter",
    )
    basis = values_for_section(
        sections,
        spec["section"],
        "No. of basis functions",
    ).astype(int)

    item = {**spec, "times": total_times, "basis": basis}
    if spec["kind"] == "pyfock":
        eri = values_for_section(sections, spec["section"], "J time (s) / iter")
        xc = values_for_section(sections, spec["section"], "XC (s) / iter")
        item["components"] = {
            "J time (s) / iter": eri,
            "XC (s) / iter": xc,
            "Other": total_times - (eri + xc),
        }

    series.append(item)

if not series:
    raise ValueError("At least one bar in bar_enabled must be set to True")

basis_functions = series[0]["basis"]

# ==== Choose x-axis ====
if x_axis_choice == "water":
    x_values = np.arange(len(water_molecules))
    x_label = "Number of Water Molecules"
    x_tick_labels = [f"(H$_2$O)$_{{{n}}}$" for n in water_molecules]
elif x_axis_choice == "basis":
    x_values = np.arange(len(basis_functions))
    x_label = "Number of Basis Functions"
    x_tick_labels = [str(n) for n in basis_functions]
else:
    raise ValueError("x_axis_choice must be 'water' or 'basis'")

# ==== Plot ====
bar_width = min(0.21, 0.94 / len(series))
offsets = (np.arange(len(series)) - (len(series) - 1) / 2) * bar_width

fig, ax = plt.subplots(figsize=(10.6, 6.8))
ax.set_axisbelow(True)
ax.grid(axis="y", color="#D5DAE0", linewidth=0.8, alpha=0.8)

for offset, item in zip(offsets, series):
    bar_positions = x_values + offset
    if item["kind"] == "pyscf":
        ax.bar(
            bar_positions,
            item["times"],
            bar_width,
            label=item["label"],
            color=item["color"],
            edgecolor=item["edgecolor"],
            linewidth=1.4,
        )
    else:
        bottom = np.zeros_like(item["times"])
        for column_name, _, color in pyfock_components:
            values = item["components"][column_name]
            ax.bar(
                bar_positions,
                values,
                bar_width,
                bottom=bottom,
                color=color,
                alpha=item["alpha"],
                edgecolor=item["edgecolor"],
                linewidth=1.05,
            )
            bottom += values

    for x, val in zip(bar_positions, item["times"]):
        ax.text(
            x,
            val * 1.01,
            f"{val:.1f}",
            ha="center",
            va="bottom",
            fontsize=12,
            fontweight="bold",
            rotation=0,
        )

    for x in bar_positions:
        ax.text(
            x,
            -0.035,
            item["short_label"],
            ha="right",
            va="top",
            fontsize=11,
            fontweight="bold",
            rotation=50,
            transform=ax.get_xaxis_transform(),
            clip_on=False,
        )

max_time = max(item["times"].max() for item in series)
if not log_scale:
    ax.set_ylim(top=max_time * 1.14)

# Labels & settings
ax.set_ylabel("Time per Iteration (s)", fontsize=16, fontweight="bold")

if log_scale:
    ax.set_yscale("log")

ax.set_title(
    "PySCF vs PyFock CPU/GPU Total Time per Iteration",
    fontsize=16,
    fontweight="bold",
)

# Set custom x-tick labels
ax.set_xticks(x_values)
ax.set_xticklabels([])

# Tick labels
ax.tick_params(axis="both", labelsize=14)
ax.tick_params(axis="x", pad=60)
for label in ax.get_xticklabels() + ax.get_yticklabels():
    label.set_fontweight("bold")

# Thicker border
for spine in ax.spines.values():
    spine.set_linewidth(1.5)

# Legend styling
component_handles = [
    Patch(facecolor="#9BB7D4", edgecolor="#1F4E79", label="PySCF total"),
]
component_handles.extend(
    Patch(facecolor=color, edgecolor="black", label=f"PyFock {label}")
    for _, label, color in pyfock_components
)

legend = ax.legend(
    handles=component_handles,
    fontsize=12,
    ncols=2,
    frameon=True,
    framealpha=0.95,
    edgecolor="#9CA3AF",
    loc="upper left",
)
for text in legend.get_texts():
    text.set_fontweight("bold")

ax.margins(x=0.08)
plt.tight_layout()
plt.show()
